import io
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse
from django.utils.translation import gettext, override
from openpyxl import load_workbook

from apps.inventory.forms import ClientForm, ProductForm
from apps.inventory.models import Brand, Category, Client, Product, ProductPackaging, Unit
from apps.inventory.pricing import get_sale_price


class CustomerPricingTests(TestCase):
    def setUp(self):
        self.category = Category.objects.create(name='Tarifs')
        self.brand = Brand.objects.create(name='Marque tarifs')
        self.unit = Unit.objects.create(name='Pièce')
        self.product = Product.objects.create(
            name='Produit trois tarifs',
            category=self.category,
            brand=self.brand,
            unit=self.unit,
            purchase_price='100.00',
            super_wholesale_price='110.00',
            wholesale_price='120.00',
            retail_price='135.00',
            sale_price='135.00',
            quantity=10,
        )

    def product_form(self, **overrides):
        data = {
            'name': 'Nouveau produit tarifé',
            'brand_text': self.brand.name,
            'purchase_price': '100.00',
            'super_wholesale_price': '110.00',
            'wholesale_price': '120.00',
            'retail_price': '135.00',
            'quantity': '5',
            'minimum_stock': '1',
            'description': '',
        }
        data.update(overrides)
        return ProductForm(data=data)

    def test_product_form_saves_three_prices_and_syncs_legacy_price(self):
        form = self.product_form()

        self.assertTrue(form.is_valid(), form.errors.as_text())
        product = form.save()
        self.assertEqual(product.super_wholesale_price, Decimal('110.00'))
        self.assertEqual(product.wholesale_price, Decimal('120.00'))
        self.assertEqual(product.retail_price, Decimal('135.00'))
        self.assertEqual(product.sale_price, product.retail_price)

    def test_each_price_must_cover_purchase_cost(self):
        cases = (
            ('super_wholesale_price', '99.00'),
            ('wholesale_price', '99.00'),
            ('retail_price', '99.00'),
        )
        for field_name, value in cases:
            with self.subTest(field_name=field_name):
                form = self.product_form(**{field_name: value})
                self.assertFalse(form.is_valid())
                self.assertIn(field_name, form.errors)

    def test_price_order_is_validated(self):
        form = self.product_form(
            super_wholesale_price='130.00',
            wholesale_price='115.00',
            retail_price='125.00',
        )

        self.assertFalse(form.is_valid())
        self.assertIn('retail_price', form.errors)

    def test_existing_client_defaults_to_retail(self):
        customer = Client.objects.create(name='Ancien client')
        self.assertEqual(customer.customer_type, Client.CustomerType.RETAIL)

    def test_authoritative_price_depends_on_customer_type(self):
        cases = (
            (Client.CustomerType.SUPER_WHOLESALE, Decimal('110.00')),
            (Client.CustomerType.WHOLESALE, Decimal('120.00')),
            (Client.CustomerType.RETAIL, Decimal('135.00')),
        )
        for customer_type, expected in cases:
            with self.subTest(customer_type=customer_type):
                customer = Client(name='Client', customer_type=customer_type)
                self.assertEqual(get_sale_price(self.product, customer), expected)

    def test_client_form_rejects_unknown_customer_type(self):
        form = ClientForm(data={'name': 'Type invalide', 'customer_type': 'UNKNOWN'})
        self.assertFalse(form.is_valid())
        self.assertIn('customer_type', form.errors)

    def test_client_detail_displays_customer_type_without_email(self):
        User = get_user_model()
        manager = User.objects.create_user(
            username='client-detail-manager', password='StrongPass123!', role=User.MANAGER
        )
        customer = Client.objects.create(
            name='Client fiche gros',
            customer_type=Client.CustomerType.WHOLESALE,
            email='hidden@example.com',
        )
        self.client.force_login(manager)

        response = self.client.get(reverse('client_detail', args=[customer.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Client fiche gros')
        self.assertContains(response, 'Gros')
        self.assertNotContains(response, 'hidden@example.com')

    def test_seller_price_endpoint_returns_only_applicable_price(self):
        User = get_user_model()
        seller = User.objects.create_user(
            username='pricing-seller', password='StrongPass123!', role=User.SELLER
        )
        self.client.force_login(seller)

        cases = (
            (Client.CustomerType.SUPER_WHOLESALE, '110.00'),
            (Client.CustomerType.WHOLESALE, '120.00'),
            (Client.CustomerType.RETAIL, '135.00'),
        )
        for customer_type, expected_price in cases:
            with self.subTest(customer_type=customer_type):
                customer = Client.objects.create(
                    name=f'Client {customer_type}', customer_type=customer_type
                )
                response = self.client.get(
                    reverse('sale_price_lookup'),
                    {'product_id': self.product.pk, 'client_id': customer.pk},
                )

                self.assertEqual(response.status_code, 200)
                payload = response.json()
                self.assertEqual(payload['product_id'], self.product.pk)
                self.assertEqual(payload['product_name'], self.product.name)
                self.assertEqual(payload['reference'], self.product.reference)
                self.assertEqual(payload['stock'], self.product.quantity)
                self.assertEqual(payload['customer_type'], customer_type)
                self.assertEqual(payload['price'], expected_price)
                self.assertNotIn('purchase_price', payload)
                self.assertNotIn('super_wholesale_price', payload)
                self.assertNotIn('wholesale_price', payload)
                self.assertNotIn('retail_price', payload)

    def test_web_price_endpoint_applies_packaging_factor(self):
        User = get_user_model()
        seller = User.objects.create_user(
            username='packaging-pricing-seller',
            password='StrongPass123!',
            role=User.SELLER,
        )
        customer = Client.objects.create(
            name='Client carton gros', customer_type=Client.CustomerType.WHOLESALE
        )
        packaging = ProductPackaging.objects.create(
            product=self.product,
            name='Carton de 6',
            conversion_factor=6,
            default_sale_price='720.00',
        )
        self.client.force_login(seller)

        response = self.client.get(reverse('sale_price_lookup'), {
            'product_id': self.product.pk,
            'client_id': customer.pk,
            'packaging_id': packaging.pk,
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['price'], '720.00')
        self.assertEqual(payload['packaging_id'], packaging.pk)
        self.assertEqual(
            payload['packagings'],
            [{
                'id': packaging.pk,
                'name': 'Carton de 6',
                'conversion_factor': 6,
                'price': '720.00',
            }],
        )

    def test_sale_price_feedback_is_translated(self):
        translations = {
            'fr': 'Impossible de récupérer le prix du produit.',
            'en': 'Unable to retrieve the product price.',
            'ar': 'تعذر استرجاع سعر المنتج.',
        }
        for language, expected in translations.items():
            with self.subTest(language=language), override(language):
                self.assertEqual(
                    gettext('Impossible de récupérer le prix du produit.'),
                    expected,
                )

    def test_customer_type_is_exported_without_email(self):
        User = get_user_model()
        manager = User.objects.create_user(
            username='pricing-manager', password='StrongPass123!', role=User.MANAGER
        )
        Client.objects.create(
            name='Client export',
            customer_type=Client.CustomerType.SUPER_WHOLESALE,
            email='private@example.com',
        )
        self.client.force_login(manager)

        response = self.client.get(reverse('client_export'))

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(io.BytesIO(response.content), read_only=True, data_only=True)
        rows = list(workbook.active.iter_rows(values_only=True))
        self.assertIn('Type de client', rows[0])
        self.assertNotIn('Email', rows[0])
        self.assertIn('Super Gros', rows[1])
        workbook.close()

    def test_product_export_keeps_import_column_contract(self):
        User = get_user_model()
        manager = User.objects.create_user(
            username='export-manager', password='StrongPass123!', role=User.MANAGER
        )
        self.client.force_login(manager)

        response = self.client.get(reverse('product_export'))

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(io.BytesIO(response.content), read_only=True, data_only=True)
        headers = list(next(workbook.active.iter_rows(values_only=True)))
        values = list(next(workbook.active.iter_rows(min_row=2, values_only=True)))
        self.assertEqual(
            headers,
            [
                'Référence', 'Code-barres', 'Nom produit', 'Catégorie', 'Marque',
                'Unité', "Prix d'achat", 'Prix de vente', 'Quantité',
                'Stock minimum', 'Prix Super Gros', 'Prix Gros', 'Prix Détail',
            ],
        )
        self.assertEqual(values[8:13], [10, 0, 110, 120, 135])
        workbook.close()
