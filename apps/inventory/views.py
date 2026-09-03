from decimal import Decimal, InvalidOperation

import openpyxl
from django.contrib import messages
from django.core.exceptions import ValidationError
from django.db import DatabaseError, transaction
from django.db.models import Exists, OuterRef, Q
from django.db.models.deletion import ProtectedError
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from apps.accounts.permissions import manager_required, seller_required
from apps.core.export_security import excel_safe_text
from apps.core.pagination import paginate_queryset
from .forms import (
    ClientForm,
    ImportExcelForm,
    ProductForm,
    ProductPackagingFormSet,
    StockMovementForm,
    SupplierForm,
)
from .models import Brand, Category, Client, Product, StockMovement, Supplier, Unit
from .services import record_stock_movement, reverse_stock_movement
from django.http import FileResponse
from django.utils.translation import gettext as _

@seller_required
def product_list(request):
    products = Product.objects.select_related('category', 'brand', 'unit')
    query = request.GET.get('q', '').strip()
    if query:
        products = products.filter(
            Q(name__icontains=query)
            | Q(reference__icontains=query)
            | Q(barcode__icontains=query)
        )
    page_obj, pagination_query = paginate_queryset(request, products.order_by('name', 'pk'))
    return render(
        request,
        'inventory/product_list.html',
        {'products': page_obj, 'page_obj': page_obj, 'pagination_query': pagination_query},
    )

@manager_required
def product_create(request):
    form = ProductForm(request.POST or None, request.FILES or None)
    packaging_data = request.POST if 'packagings-TOTAL_FORMS' in request.POST else None
    packaging_formset = ProductPackagingFormSet(packaging_data, prefix='packagings')
    if form.is_valid() and (not packaging_formset.is_bound or packaging_formset.is_valid()):
        try:
            with transaction.atomic():
                initial_quantity = form.cleaned_data['quantity']
                product = form.save(commit=False)
                product.quantity = 0
                product.save()
                if packaging_formset.is_bound:
                    packaging_formset.instance = product
                    packaging_formset.save()
                if initial_quantity:
                    record_stock_movement(
                        product=product,
                        movement_type=StockMovement.ENTRY,
                        quantity=initial_quantity,
                        reason=_('Stock initial du produit'),
                        user=request.user,
                        source_type=StockMovement.SOURCE_PRODUCT,
                        source_reference=product.reference,
                    )
        except ValidationError as exc:
            for error in exc.messages:
                form.add_error('quantity', error)
        else:
            messages.success(request, _('Produit ajouté avec succès.'))
            if getattr(product, '_generated_media_errors', None):
                messages.warning(
                    request,
                    _(
                        'Le produit a été créé, mais ses images QR/code-barres '
                        'n’ont pas pu être stockées. Vérifiez le stockage média.'
                    ),
                )
            return redirect('product_list')
    return render(request, 'inventory/product_form.html', {
        'form': form, 'packaging_formset': packaging_formset, 'title': _('Ajouter un produit'),
    })

@manager_required
def product_update(request, pk):
    product = get_object_or_404(Product, pk=pk)
    form = ProductForm(request.POST or None, request.FILES or None, instance=product)
    packaging_data = request.POST if 'packagings-TOTAL_FORMS' in request.POST else None
    packaging_formset = ProductPackagingFormSet(packaging_data, instance=product, prefix='packagings')
    if form.is_valid() and (not packaging_formset.is_bound or packaging_formset.is_valid()):
        try:
            with transaction.atomic():
                target_quantity = form.cleaned_data['quantity']
                locked_product = Product.objects.select_for_update().get(pk=product.pk)
                form.instance.quantity = locked_product.quantity
                saved_product = form.save()
                if packaging_formset.is_bound:
                    packaging_formset.instance = saved_product
                    packaging_formset.save()
                if target_quantity != locked_product.quantity:
                    record_stock_movement(
                        product=saved_product,
                        movement_type=StockMovement.ADJUSTMENT,
                        quantity=target_quantity,
                        reason=_('Modification du stock depuis la fiche produit'),
                        user=request.user,
                        source_type=StockMovement.SOURCE_PRODUCT,
                        source_reference=saved_product.reference,
                    )
        except ValidationError as exc:
            for error in exc.messages:
                form.add_error('quantity', error)
        else:
            messages.success(request, _('Produit mis à jour.'))
            return redirect('product_list')
    return render(request, 'inventory/product_form.html', {
        'form': form, 'packaging_formset': packaging_formset, 'title': _('Modifier le produit'),
    })

@manager_required
def product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        try:
            product.delete()
        except ProtectedError:
            messages.error(
                request,
                _('Ce produit ne peut pas être supprimé car il possède un historique de stock ou commercial.'),
            )
        else:
            messages.success(request, _('Produit supprimé.'))
        return redirect('product_list')
    return render(request, 'inventory/product_confirm_delete.html', {'product': product})

@manager_required
def product_import(request):
    form = ImportExcelForm(request.POST or None, request.FILES or None)
    if form.is_valid():
        try:
            with transaction.atomic():
                workbook = openpyxl.load_workbook(form.cleaned_data['file'], read_only=True, data_only=True)
                sheet = workbook.active
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    reference, barcode, name, category, brand, unit, purchase_price, sale_price, quantity, minimum_stock = row[:10]
                    tier_values = list(row[10:13])
                    tier_values.extend([None] * (3 - len(tier_values)))
                    reference = str(reference or '').strip()
                    name = str(name or '').strip()
                    if not name:
                        raise ValueError('missing product name')

                    try:
                        purchase_price = Decimal(str(purchase_price or 0))
                        sale_price = Decimal(str(sale_price or 0))
                        super_wholesale_price = Decimal(str(tier_values[0] if tier_values[0] is not None else sale_price))
                        wholesale_price = Decimal(str(tier_values[1] if tier_values[1] is not None else sale_price))
                        retail_price = Decimal(str(tier_values[2] if tier_values[2] is not None else sale_price))
                        quantity_decimal = Decimal(str(quantity or 0))
                        minimum_stock_decimal = Decimal(str(minimum_stock or 0))
                    except InvalidOperation as exc:
                        raise ValueError('invalid numeric value') from exc
                    if (
                        purchase_price < 0
                        or sale_price < 0
                        or super_wholesale_price < purchase_price
                        or wholesale_price < purchase_price
                        or retail_price < purchase_price
                        or not super_wholesale_price <= wholesale_price <= retail_price
                        or quantity_decimal < 0
                        or minimum_stock_decimal < 0
                        or quantity_decimal != quantity_decimal.to_integral_value()
                        or minimum_stock_decimal != minimum_stock_decimal.to_integral_value()
                    ):
                        raise ValueError('invalid stock or price value')

                    category_obj = Category.objects.resolve(category) if category else None
                    brand_obj = Brand.objects.resolve(brand) if brand else None
                    unit_obj = Unit.objects.resolve(unit) if unit else None
                    product = None
                    if reference:
                        product = Product.objects.select_for_update().filter(reference=reference).first()

                    target_quantity = int(quantity_decimal)
                    if product is None:
                        product = Product(
                            reference=reference,
                            barcode=str(barcode or '').strip(),
                            name=name,
                            category=category_obj,
                            brand=brand_obj,
                            unit=unit_obj,
                            purchase_price=purchase_price,
                            sale_price=sale_price,
                            super_wholesale_price=super_wholesale_price,
                            wholesale_price=wholesale_price,
                            retail_price=retail_price,
                            quantity=0,
                            minimum_stock=int(minimum_stock_decimal),
                        )
                        product.save()
                    else:
                        current_quantity = product.quantity
                        product.barcode = str(barcode or '').strip()
                        product.name = name
                        product.category = category_obj
                        product.brand = brand_obj
                        product.unit = unit_obj
                        product.purchase_price = purchase_price
                        product.sale_price = sale_price
                        product.super_wholesale_price = super_wholesale_price
                        product.wholesale_price = wholesale_price
                        product.retail_price = retail_price
                        product.quantity = current_quantity
                        product.minimum_stock = int(minimum_stock_decimal)
                        product.save()

                    if product.quantity != target_quantity:
                        record_stock_movement(
                            product=product,
                            movement_type=StockMovement.ADJUSTMENT,
                            quantity=target_quantity,
                            reason=_('Synchronisation du stock par import Excel'),
                            user=request.user,
                            source_type=StockMovement.SOURCE_IMPORT,
                            source_reference=product.reference,
                        )
                workbook.close()
        except (
            DatabaseError,
            OSError,
            TypeError,
            ValueError,
            ValidationError,
            openpyxl.utils.exceptions.InvalidFileException,
        ):
            form.add_error('file', _('Le fichier Excel contient des données invalides. Aucune donnée n’a été importée.'))
        else:
            messages.success(request, _('Importation Excel terminée.'))
            return redirect('product_list')
    return render(request, 'inventory/product_import.html', {'form': form})

@manager_required
def product_export(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = str(_('Produits'))[:31]
    headers = [
        _('Référence'),
        _('Code-barres'),
        _('Nom produit'),
        _('Catégorie'),
        _('Marque'),
        _('Unité'),
        _("Prix d'achat"),
        _('Prix de vente'),
        _('Quantité'),
        _('Stock minimum'),
        _('Prix Super Gros'),
        _('Prix Gros'),
        _('Prix Détail'),
    ]
    sheet.append(headers)
    for product in Product.objects.all():
        sheet.append([
            excel_safe_text(product.reference),
            excel_safe_text(product.barcode),
            excel_safe_text(product.name),
            excel_safe_text(product.category),
            excel_safe_text(product.brand),
            excel_safe_text(product.unit),
            float(product.purchase_price),
            float(product.sale_price),
            product.quantity,
            product.minimum_stock,
            float(product.super_wholesale_price),
            float(product.wholesale_price),
            float(product.retail_price),
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=produits.xlsx'
    workbook.save(response)
    return response

@manager_required
def client_list(request):
    clients = Client.objects.order_by('name', 'pk')
    query = request.GET.get('q', '').strip()
    customer_type = request.GET.get('customer_type', '').strip()
    if query:
        clients = clients.filter(
            Q(name__icontains=query)
            | Q(phone__icontains=query)
            | Q(email__icontains=query)
            | Q(tax_number__icontains=query)
        )
    if customer_type in Client.CustomerType.values:
        clients = clients.filter(customer_type=customer_type)
    page_obj, pagination_query = paginate_queryset(request, clients)
    return render(
        request,
        'inventory/client_list.html',
        {
            'clients': page_obj,
            'page_obj': page_obj,
            'pagination_query': pagination_query,
            'customer_types': Client.CustomerType.choices,
            'selected_customer_type': customer_type,
        },
    )


@manager_required
def client_export(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = str(_('Clients'))[:31]
    sheet.append([
        _('Nom'),
        _('Téléphone'),
        _('Adresse'),
        _('Wilaya'),
        _('Type de client'),
        _('NIF'),
        _('Solde'),
        _('Notes'),
    ])
    for client in Client.objects.order_by('name', 'pk'):
        sheet.append([
            excel_safe_text(client.name),
            excel_safe_text(client.phone),
            excel_safe_text(client.address),
            excel_safe_text(client.wilaya),
            excel_safe_text(client.get_customer_type_display()),
            excel_safe_text(client.tax_number),
            float(client.balance),
            excel_safe_text(client.notes),
        ])
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=clients.xlsx'
    workbook.save(response)
    return response


@seller_required
def client_create(request):
    form = ClientForm(request.POST or None)
    if form.is_valid():
        form.save()
        messages.success(request, _('Client ajouté avec succès.'))
        return redirect('client_list')
    return render(request, 'inventory/client_form.html', {'form': form, 'title': _('Ajouter un client')})


@manager_required
def client_detail(request, pk):
    client_obj = get_object_or_404(Client, pk=pk)
    sales = client_obj.sales.prefetch_related('lines__product').order_by('-created_at', '-pk')[:50]
    return render(
        request,
        'inventory/client_detail.html',
        {'client_obj': client_obj, 'sales': sales},
    )


@manager_required
def client_update(request, pk):
    client_obj = get_object_or_404(Client, pk=pk)
    form = ClientForm(request.POST or None, instance=client_obj)
    if form.is_valid():
        form.save()
        messages.success(request, _('Client mis à jour.'))
        return redirect('client_list')
    return render(request, 'inventory/client_form.html', {'form': form, 'title': _('Modifier le client')})

@manager_required
def client_delete(request, pk):
    client_obj = get_object_or_404(Client, pk=pk)
    if request.method == 'POST':
        client_obj.delete()
        messages.success(request, _('Client supprimé.'))
        return redirect('client_list')
    return render(request, 'inventory/client_confirm_delete.html', {'client': client_obj})

@manager_required
def supplier_list(request):
    suppliers = Supplier.objects.order_by('name', 'pk')
    query = request.GET.get('q', '').strip()
    if query:
        suppliers = suppliers.filter(
            Q(name__icontains=query)
            | Q(phone__icontains=query)
            | Q(email__icontains=query)
            | Q(rc_number__icontains=query)
            | Q(tax_number__icontains=query)
        )
    page_obj, pagination_query = paginate_queryset(request, suppliers)
    return render(
        request,
        'inventory/supplier_list.html',
        {'suppliers': page_obj, 'page_obj': page_obj, 'pagination_query': pagination_query},
    )

@manager_required
def supplier_create(request):
    form = SupplierForm(request.POST or None)
    if form.is_valid():
        form.save()
        messages.success(request, _('Fournisseur ajouté avec succès.'))
        return redirect('supplier_list')
    return render(request, 'inventory/supplier_form.html', {'form': form, 'title': _('Ajouter un fournisseur')})

@manager_required
def supplier_update(request, pk):
    supplier_obj = get_object_or_404(Supplier, pk=pk)
    form = SupplierForm(request.POST or None, instance=supplier_obj)
    if form.is_valid():
        form.save()
        messages.success(request, _('Fournisseur mis à jour.'))
        return redirect('supplier_list')
    return render(request, 'inventory/supplier_form.html', {'form': form, 'title': _('Modifier le fournisseur')})

@manager_required
def supplier_delete(request, pk):
    supplier_obj = get_object_or_404(Supplier, pk=pk)
    if request.method == 'POST':
        supplier_obj.delete()
        messages.success(request, _('Fournisseur supprimé.'))
        return redirect('supplier_list')
    return render(request, 'inventory/supplier_confirm_delete.html', {'supplier': supplier_obj})

@seller_required
def stock_movement_list(request):
    movements = (
        StockMovement.objects.select_related('product', 'created_by')
        .annotate(has_reversal=Exists(StockMovement.objects.filter(reversal_of=OuterRef('pk'))))
        .order_by('-created_at', '-pk')
    )
    page_obj, pagination_query = paginate_queryset(request, movements)
    return render(
        request,
        'inventory/stock_movement_list.html',
        {'movements': page_obj, 'page_obj': page_obj, 'pagination_query': pagination_query},
    )

@manager_required
def stock_movement_create(request):
    form = StockMovementForm(request.POST or None)
    if form.is_valid():
        movement = form.save(commit=False)
        try:
            record_stock_movement(
                product=movement.product,
                movement_type=movement.movement_type,
                quantity=movement.quantity,
                reason=movement.reason,
                user=request.user,
            )
        except ValidationError as exc:
            if hasattr(exc, 'message_dict'):
                for field, errors in exc.message_dict.items():
                    target = field if field in form.fields else None
                    for error in errors:
                        form.add_error(target, error)
            else:
                for error in exc.messages:
                    form.add_error(None, error)
        else:
            messages.success(request, _('Mouvement de stock enregistré.'))
            return redirect('stock_movement_list')
    return render(request, 'inventory/stock_movement_form.html', {'form': form, 'title': _('Ajouter un mouvement de stock')})

@manager_required
def stock_movement_delete(request, pk):
    movement = get_object_or_404(StockMovement, pk=pk)
    if request.method == 'POST':
        try:
            reverse_stock_movement(movement, user=request.user)
        except ValidationError as exc:
            messages.error(request, ' '.join(exc.messages))
        else:
            messages.success(request, _('Mouvement annulé par une écriture de contrepassation.'))
        return redirect('stock_movement_list')
    return render(request, 'inventory/stock_movement_confirm_delete.html', {'movement': movement})


@seller_required
def product_detail(request, pk):
    product = get_object_or_404(Product.objects.prefetch_related('packagings'), pk=pk)
    if not product.barcode or not product.barcode_image or not product.qr_code:
        product.save()
        product.refresh_from_db()
    return render(request, 'inventory/product_detail.html', {'product': product})


@seller_required
def product_qr_download(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if not product.qr_code:
        messages.error(request, _('Aucun QR code disponible pour ce produit.'))
        return redirect('product_detail', pk=pk)
    return FileResponse(product.qr_code.open('rb'), as_attachment=True, filename=f"{product.reference}_qr.png")


@seller_required
def product_barcode_download(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if not product.barcode_image:
        product.save()
        product.refresh_from_db()
    if not product.barcode_image:
        messages.error(request, _('Aucun code-barres disponible pour ce produit.'))
        return redirect('product_detail', pk=pk)
    return FileResponse(product.barcode_image.open('rb'), as_attachment=True, filename=f"{product.reference}_barcode.svg")
