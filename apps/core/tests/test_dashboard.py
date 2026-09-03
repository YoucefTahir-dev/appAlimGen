import io
from decimal import Decimal
from datetime import date, datetime, time, timedelta
from unittest.mock import patch

import openpyxl
from django.test import RequestFactory, TestCase
from django.urls import reverse
from django.contrib.auth import get_user_model
from django.utils import timezone
from apps.inventory.models import Product, Category, Brand, Unit, Client, Supplier
from apps.commerce.models import Purchase, PurchaseLine, Sale, SaleLine
from apps.expenses.models import Expense, ExpenseCategory

from apps.core.dashboard import get_period_bounds


class DashboardTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(
            username='tester',
            password='pass',
            role=User.ADMIN,
            is_staff=True,
            is_superuser=True,
        )
        self.client.login(username='tester', password='pass')
        cat = Category.objects.create(name='Cat1')
        brand = Brand.objects.create(name='Brand1')
        unit = Unit.objects.create(name='Unit1')
        self.product = Product.objects.create(
            barcode='123',
            name='Test Product',
            category=cat,
            brand=brand,
            unit=unit,
            purchase_price='10.00',
            sale_price='15.00',
            quantity=10,
            minimum_stock=1,
        )
        self.client_obj = Client.objects.create(name='Client1')
        self.supplier_obj = Supplier.objects.create(name='Supplier1')
        self.sale = Sale.objects.create(invoice_number='INV123', client=self.client_obj, total='100', discount='0', tax_rate='10')
        self.purchase = Purchase.objects.create(reference='PO123', supplier=self.supplier_obj, total='50', tax_rate='10')
        SaleLine.objects.create(sale=self.sale, product=self.product, quantity=2, unit_price='50.00')
        PurchaseLine.objects.create(purchase=self.purchase, product=self.product, quantity=3, purchase_price='10.00')
        self.expense_category, _ = ExpenseCategory.objects.get_or_create(name='Divers')
        Expense.objects.create(
            category=self.expense_category,
            description='Charge test',
            amount='15.00',
            created_by=self.user,
            date=timezone.localdate(),
        )

    @staticmethod
    def local_noon(day):
        return timezone.make_aware(
            datetime.combine(day, time(hour=12)),
            timezone.get_current_timezone(),
        )

    def move_base_records_to(self, day):
        Sale.objects.filter(pk=self.sale.pk).update(created_at=self.local_noon(day))
        Purchase.objects.filter(pk=self.purchase.pk).update(created_at=self.local_noon(day))
        Expense.objects.filter(description='Charge test').update(date=day)

    def test_dashboard_loads(self):
        response = self.client.get(reverse('dashboard'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Tableau de bord')
        self.assertContains(response, 'Valeur du stock')
        self.assertContains(response, "Chiffre d'affaires de la période")
        self.assertContains(response, "Gain brut (avant charges)")
        self.assertContains(response, "Gain net (après charges)")
        self.assertContains(response, "65,00")

    def test_historical_profit_uses_the_sale_cost_snapshot(self):
        self.product.purchase_price = Decimal('90.00')
        self.product.super_wholesale_price = Decimal('100.00')
        self.product.wholesale_price = Decimal('100.00')
        self.product.retail_price = Decimal('100.00')
        self.product.save(
            update_fields=[
                'purchase_price',
                'super_wholesale_price',
                'wholesale_price',
                'retail_price',
            ]
        )

        response = self.client.get(reverse('dashboard'), {'period': 'today'})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['gross_profit'], Decimal('80.00'))
        self.assertEqual(response.context['net_profit'], Decimal('65.00'))

    def test_unreconciled_legacy_documents_do_not_create_false_unpaid_alerts(self):
        self.sale.payment_tracking_initialized = False
        self.sale.save(update_fields=['payment_tracking_initialized'])
        self.purchase.payment_tracking_initialized = False
        self.purchase.save(update_fields=['payment_tracking_initialized'])

        response = self.client.get(reverse('dashboard'), {'period': 'today'})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['unpaid_invoices'], 0)
        self.assertEqual(response.context['pending_supplier_payments'], 0)

    def test_dashboard_loads_in_arabic_rtl(self):
        self.client.cookies['django_language'] = 'ar'

        response = self.client.get(reverse('dashboard'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'dir="rtl"')
        self.assertContains(response, 'لوحة التحكم')
        self.assertContains(response, 'قيمة المخزون')

    def test_mobile_sidebar_uses_bootstrap_offcanvas(self):
        response = self.client.get(reverse('dashboard'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'bootstrap.bundle.min.js')
        self.assertContains(response, 'id="appSidebar"')
        self.assertContains(response, 'offcanvas-lg offcanvas-start')
        self.assertContains(response, 'data-bs-toggle="offcanvas"')
        self.assertContains(response, 'data-bs-target="#appSidebar"')
        self.assertContains(response, 'data-bs-dismiss="offcanvas"')
        self.assertNotContains(response, 'sidebar-open')

    def test_dashboard_period_filter_and_exports(self):
        response = self.client.get(reverse('dashboard'), {'period': 'today'})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '100,00')
        self.assertContains(response, '50,00')
        self.assertContains(response, '15,00')
        self.assertContains(response, '80,00')

        excel_response = self.client.get(reverse('dashboard_export_excel'), {'period': 'today'})
        self.assertEqual(excel_response.status_code, 200)
        self.assertIn('spreadsheetml', excel_response['Content-Type'])

        pdf_response = self.client.get(reverse('dashboard_export_pdf'), {'period': 'today'})
        self.assertEqual(pdf_response.status_code, 200)
        self.assertEqual(pdf_response['Content-Type'], 'application/pdf')

    def test_predefined_periods_use_matching_calendar_comparisons(self):
        request_factory = RequestFactory()
        frozen_today = date(2026, 8, 15)
        expected = {
            'today': (date(2026, 8, 15), date(2026, 8, 15), date(2026, 8, 14), date(2026, 8, 14)),
            'yesterday': (date(2026, 8, 14), date(2026, 8, 14), date(2026, 8, 13), date(2026, 8, 13)),
            'week': (date(2026, 8, 10), date(2026, 8, 15), date(2026, 8, 3), date(2026, 8, 8)),
            'month': (date(2026, 8, 1), date(2026, 8, 15), date(2026, 7, 1), date(2026, 7, 15)),
            'year': (date(2026, 1, 1), date(2026, 8, 15), date(2025, 1, 1), date(2025, 8, 15)),
        }

        with patch('apps.core.dashboard.timezone.localdate', return_value=frozen_today):
            for period, dates in expected.items():
                with self.subTest(period=period):
                    bounds = get_period_bounds(request_factory.get('/', {'period': period}))
                    self.assertEqual(
                        (
                            bounds['start_date'],
                            bounds['end_date'],
                            bounds['previous_start_date'],
                            bounds['previous_end_date'],
                        ),
                        dates,
                    )

    def test_custom_period_filters_every_decision_indicator_and_top(self):
        selected_day = timezone.localdate() - timedelta(days=10)
        outside_day = selected_day - timedelta(days=10)
        self.move_base_records_to(selected_day)

        other_client = Client.objects.create(name='Client hors période')
        other_supplier = Supplier.objects.create(name='Fournisseur hors période')
        outside_sale = Sale.objects.create(
            invoice_number='INV-OUTSIDE',
            client=other_client,
            total='300.00',
            discount='0.00',
            tax_rate='0.00',
            created_at=self.local_noon(outside_day),
        )
        SaleLine.objects.create(sale=outside_sale, product=self.product, quantity=1, unit_price='300.00')
        outside_purchase = Purchase.objects.create(
            reference='PO-OUTSIDE',
            supplier=other_supplier,
            total='250.00',
            tax_rate='0.00',
            created_at=self.local_noon(outside_day),
        )
        PurchaseLine.objects.create(
            purchase=outside_purchase,
            product=self.product,
            quantity=1,
            purchase_price='10.00',
        )
        Expense.objects.create(
            category=self.expense_category,
            description='Charge hors période',
            amount='75.00',
            created_by=self.user,
            date=outside_day,
        )

        response = self.client.get(reverse('dashboard'), {
            'period': 'custom',
            'start_date': selected_day.isoformat(),
            'end_date': selected_day.isoformat(),
        })

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['period_revenue'], 100)
        self.assertEqual(response.context['purchases_total'], 50)
        self.assertEqual(response.context['expenses_total'], 15)
        self.assertEqual(response.context['sales_count'], 1)
        self.assertEqual(response.context['average_basket'], 100)
        self.assertEqual(response.context['products_sold'], 2)
        self.assertEqual(response.context['products_purchased'], 3)
        self.assertEqual(response.context['gross_profit'], 80)
        self.assertEqual(response.context['net_profit'], 65)
        self.assertEqual(response.context['top_products'][0]['product__name'], self.product.name)
        self.assertEqual(response.context['top_products'][0]['quantity'], 2)
        self.assertEqual(response.context['top_clients'][0]['client__name'], self.client_obj.name)
        self.assertEqual(response.context['top_suppliers'][0]['supplier__name'], self.supplier_obj.name)
        self.assertEqual(response.context['profitable_products'][0]['profit'], 80)

    def test_custom_period_chart_contains_zero_value_days(self):
        middle_day = timezone.localdate() - timedelta(days=10)
        self.move_base_records_to(middle_day)
        start_day = middle_day - timedelta(days=1)
        end_day = middle_day + timedelta(days=1)

        response = self.client.get(reverse('dashboard'), {
            'period': 'custom',
            'start_date': start_day.isoformat(),
            'end_date': end_day.isoformat(),
        })

        trend = response.context['chart_data']['trend']
        self.assertEqual(
            trend['labels'],
            [start_day.isoformat(), middle_day.isoformat(), end_day.isoformat()],
        )
        self.assertEqual(trend['revenue'], [0, 100.0, 0])
        self.assertEqual(trend['purchases'], [0, 50.0, 0])
        self.assertEqual(trend['expenses'], [0, 15.0, 0])

    def test_invalid_custom_period_is_reported_and_exports_refuse_it(self):
        today = timezone.localdate()
        invalid_params = {
            'period': 'custom',
            'start_date': today.isoformat(),
            'end_date': (today - timedelta(days=1)).isoformat(),
        }

        response = self.client.get(reverse('dashboard'), invalid_params)

        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.context['period_is_valid'])
        self.assertContains(response, 'La période demandée est invalide.')
        self.assertContains(response, 'La date de fin doit être postérieure ou égale')
        self.assertNotContains(response, reverse('dashboard_export_pdf') + '?')
        self.assertEqual(self.client.get(reverse('dashboard_export_excel'), invalid_params).status_code, 400)
        self.assertEqual(self.client.get(reverse('dashboard_export_pdf'), invalid_params).status_code, 400)

    def test_datetime_filter_uses_local_half_open_day_boundaries(self):
        selected_day = timezone.localdate() - timedelta(days=5)
        start_dt = timezone.make_aware(
            datetime.combine(selected_day, time.min),
            timezone.get_current_timezone(),
        )
        next_day_dt = timezone.make_aware(
            datetime.combine(selected_day + timedelta(days=1), time.min),
            timezone.get_current_timezone(),
        )
        Sale.objects.filter(pk=self.sale.pk).update(created_at=start_dt)
        excluded_sale = Sale.objects.create(
            invoice_number='INV-NEXT-DAY',
            client=self.client_obj,
            total='900.00',
            discount='0.00',
            tax_rate='0.00',
            created_at=next_day_dt,
        )
        SaleLine.objects.create(sale=excluded_sale, product=self.product, quantity=1, unit_price='900.00')

        response = self.client.get(reverse('dashboard'), {
            'period': 'custom',
            'start_date': selected_day.isoformat(),
            'end_date': selected_day.isoformat(),
        })

        self.assertEqual(response.context['period_revenue'], 100)
        self.assertEqual(response.context['sales_count'], 1)

    def test_exports_keep_selected_period_and_embed_charts(self):
        selected_day = timezone.localdate() - timedelta(days=7)
        self.move_base_records_to(selected_day)
        params = {
            'period': 'custom',
            'start_date': selected_day.isoformat(),
            'end_date': selected_day.isoformat(),
        }

        excel_response = self.client.get(reverse('dashboard_export_excel'), params)

        self.assertEqual(excel_response.status_code, 200)
        workbook = openpyxl.load_workbook(io.BytesIO(excel_response.content))
        dashboard_sheet = workbook['Tableau de bord']
        self.assertEqual(dashboard_sheet.cell(row=1, column=2).value, selected_day.isoformat())
        self.assertEqual(dashboard_sheet.cell(row=1, column=3).value, selected_day.isoformat())
        exported_indicators = {
            row[0].value: row[1].value
            for row in dashboard_sheet.iter_rows()
            if row and row[0].value and len(row) > 1
        }
        self.assertEqual(exported_indicators["Chiffre d'affaires de la période"], 100)
        self.assertEqual(exported_indicators['Achats de la période'], 50)
        self.assertEqual(exported_indicators['Total des charges'], 15)
        self.assertTrue(dashboard_sheet._charts)
        self.assertTrue(workbook['Répartitions']._charts)

        pdf_response = self.client.get(reverse('dashboard_export_pdf'), params)
        self.assertEqual(pdf_response.status_code, 200)
        self.assertTrue(pdf_response.content.startswith(b'%PDF'))
