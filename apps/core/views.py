import io

import openpyxl
from django.db import DatabaseError, connection
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.utils.translation import gettext as _
from openpyxl.chart import LineChart as ExcelLineChart
from openpyxl.chart import PieChart as ExcelPieChart
from openpyxl.chart import Reference
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.shapes import Drawing, Rect, String
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from apps.accounts.permissions import seller_required
from .dashboard import DashboardPeriodError, dashboard_context


def health(request):
    response = JsonResponse({'status': 'ok'})
    response['Cache-Control'] = 'no-store'
    return response


def readiness(request):
    try:
        with connection.cursor() as cursor:
            cursor.execute('SELECT 1')
            cursor.fetchone()
    except DatabaseError:
        response = JsonResponse({'status': 'unavailable'}, status=503)
    else:
        response = JsonResponse({'status': 'ready'})
    response['Cache-Control'] = 'no-store'
    return response


def money(value):
    return f"{value:,.2f} {_('DZD')}".replace(",", " ")


def invalid_period_response(error):
    message = _("Période invalide : %(errors)s") % {"errors": " ".join(error.messages)}
    return HttpResponse(message, status=400, content_type="text/plain; charset=utf-8")


def pdf_trend_chart(chart_data):
    trend = chart_data["trend"]
    if not trend["labels"]:
        return None

    drawing = Drawing(440, 235)
    chart = HorizontalLineChart()
    chart.x = 45
    chart.y = 50
    chart.height = 150
    chart.width = 365
    chart.data = [
        trend["revenue"],
        trend["purchases"],
        trend["expenses"],
        trend["gross_profit"],
        trend["net_profit"],
    ]
    chart.categoryAxis.categoryNames = trend["labels"]
    chart.categoryAxis.labels.fontSize = 6
    chart.categoryAxis.labels.angle = 30
    chart.valueAxis.labels.fontSize = 7
    chart.lines[0].strokeColor = colors.HexColor("#198754")
    chart.lines[1].strokeColor = colors.HexColor("#0d6efd")
    chart.lines[2].strokeColor = colors.HexColor("#dc3545")
    chart.lines[3].strokeColor = colors.HexColor("#ffc107")
    chart.lines[4].strokeColor = colors.HexColor("#6f42c1")
    drawing.add(chart)

    legend_items = [
        (_("CA"), "#198754"),
        (_("Achats"), "#0d6efd"),
        (_("Charges"), "#dc3545"),
        (_("Gain brut"), "#ffc107"),
        (_("Gain net"), "#6f42c1"),
    ]
    x_position = 45
    for label, colour in legend_items:
        drawing.add(Rect(x_position, 18, 7, 7, fillColor=colors.HexColor(colour), strokeColor=None))
        drawing.add(String(x_position + 10, 18, str(label), fontSize=7))
        x_position += 75
    return drawing


def pdf_pie_chart(title, labels, values):
    if not values:
        return None

    drawing = Drawing(210, 190)
    drawing.add(String(10, 174, title, fontSize=9))
    pie = Pie()
    pie.x = 55
    pie.y = 45
    pie.width = 120
    pie.height = 120
    pie.data = values
    pie.labels = labels
    pie.slices.fontSize = 6
    drawing.add(pie)
    return drawing


@seller_required
def dashboard(request):
    return render(request, 'core/dashboard.html', dashboard_context(request))


@seller_required
def dashboard_export_excel(request):
    try:
        context = dashboard_context(request, strict_period=True)
    except DashboardPeriodError as error:
        return invalid_period_response(error)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = str(_("Tableau de bord"))[:31]
    sheet.append([_("Période"), context["start_date"].isoformat(), context["end_date"].isoformat()])
    sheet.append([])
    sheet.append([_("Indicateur"), _("Valeur")])
    for label, value in [
        (_("Chiffre d'affaires du jour"), context["sales_today"]),
        (_("Chiffre d'affaires de la période"), context["period_revenue"]),
        (_("Nombre de ventes"), context["sales_count"]),
        (_("Panier moyen"), context["average_basket"]),
        (_("Gain brut (avant charges)"), context["gross_profit"]),
        (_("Total des charges"), context["expenses_total"]),
        (_("Gain net (après charges)"), context["net_profit"]),
        (_("Valeur du stock (instantané actuel)"), context["stock_value"]),
        (_("Nombre de produits (instantané actuel)"), context["total_products"]),
        (_("Nombre de clients (instantané actuel)"), context["total_clients"]),
        (_("Nombre de fournisseurs (instantané actuel)"), context["total_suppliers"]),
        (_("Achats de la période"), context["purchases_total"]),
        (_("Ventes de la période"), context["period_sales"]),
        (_("Produits vendus"), context["products_sold"]),
        (_("Produits achetés"), context["products_purchased"]),
        (_("Notifications"), context["notification_count"]),
    ]:
        sheet.append([label, value])

    sheet.append([])
    trend_header_row = sheet.max_row + 1
    sheet.append([_("Évolution"), _("CA"), _("Achats"), _("Charges"), _("Gain brut"), _("Gain net")])
    trend = context["chart_data"]["trend"]
    for index, label in enumerate(trend["labels"]):
        sheet.append([
            label,
            trend["revenue"][index],
            trend["purchases"][index],
            trend["expenses"][index],
            trend["gross_profit"][index],
            trend["net_profit"][index],
        ])

    if trend["labels"]:
        chart = ExcelLineChart()
        chart.title = _("Évolution financière")
        chart.y_axis.title = _("DZD")
        chart.x_axis.title = _("Période")
        chart.height = 9
        chart.width = 18
        data = Reference(
            sheet,
            min_col=2,
            max_col=6,
            min_row=trend_header_row,
            max_row=sheet.max_row,
        )
        categories = Reference(sheet, min_col=1, min_row=trend_header_row + 1, max_row=sheet.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        sheet.add_chart(chart, "H2")

    category_sheet = workbook.create_sheet(_("Répartitions")[:31])
    category_sheet.append([_("Catégorie de vente"), _("Total")])
    sales_categories = context["chart_data"]["sales_categories"]
    for label, value in zip(sales_categories["labels"], sales_categories["values"]):
        category_sheet.append([label, value])
    sales_end_row = category_sheet.max_row

    category_sheet.append([])
    expense_header_row = category_sheet.max_row + 1
    category_sheet.append([_("Catégorie de charge"), _("Total")])
    expense_categories = context["chart_data"]["expense_categories"]
    for label, value in zip(expense_categories["labels"], expense_categories["values"]):
        category_sheet.append([label, value])
    expense_end_row = category_sheet.max_row

    if sales_categories["values"]:
        sales_pie = ExcelPieChart()
        sales_pie.title = _("Ventes par catégorie")
        sales_pie.add_data(Reference(category_sheet, min_col=2, min_row=1, max_row=sales_end_row), titles_from_data=True)
        sales_pie.set_categories(Reference(category_sheet, min_col=1, min_row=2, max_row=sales_end_row))
        category_sheet.add_chart(sales_pie, "D2")
    if expense_categories["values"]:
        expense_pie = ExcelPieChart()
        expense_pie.title = _("Charges par catégorie")
        expense_pie.add_data(
            Reference(category_sheet, min_col=2, min_row=expense_header_row, max_row=expense_end_row),
            titles_from_data=True,
        )
        expense_pie.set_categories(
            Reference(category_sheet, min_col=1, min_row=expense_header_row + 1, max_row=expense_end_row)
        )
        category_sheet.add_chart(expense_pie, "D18")

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=dashboard.xlsx"
    return response


@seller_required
def dashboard_export_pdf(request):
    try:
        context = dashboard_context(request, strict_period=True)
    except DashboardPeriodError as error:
        return invalid_period_response(error)

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = "attachment; filename=dashboard.pdf"
    doc = SimpleDocTemplate(response, pagesize=A4, title=str(_("Tableau de bord")))
    styles = getSampleStyleSheet()
    story = [
        Paragraph(_("Tableau de bord décisionnel"), styles["Title"]),
        Paragraph(f"{context['start_date']:%d/%m/%Y} - {context['end_date']:%d/%m/%Y}", styles["Normal"]),
        Spacer(1, 12),
    ]
    indicators = [
        [_("Chiffre d'affaires de la période"), money(context["period_revenue"])],
        [_("Nombre de ventes"), context["sales_count"]],
        [_("Panier moyen"), money(context["average_basket"])],
        [_("Gain brut (avant charges)"), money(context["gross_profit"])],
        [_("Total des charges"), money(context["expenses_total"])],
        [_("Gain net (après charges)"), money(context["net_profit"])],
        [_("Valeur du stock (instantané actuel)"), money(context["stock_value"])],
        [_("Achats de la période"), money(context["purchases_total"])],
        [_("Produits vendus"), context["products_sold"]],
        [_("Produits achetés"), context["products_purchased"]],
        [_("Notifications"), context["notification_count"]],
    ]
    table = Table([[_("Indicateur"), _("Valeur")], *indicators], colWidths=[260, 180])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a2a")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
        ("PADDING", (0, 0), (-1, -1), 6),
    ]))
    story.extend([table, Spacer(1, 12), Paragraph(_("Graphiques de la période"), styles["Heading2"])])
    trend_drawing = pdf_trend_chart(context["chart_data"])
    if trend_drawing is not None:
        story.extend([trend_drawing, Spacer(1, 8)])

    sales_categories = context["chart_data"]["sales_categories"]
    expense_categories = context["chart_data"]["expense_categories"]
    category_drawings = [
        pdf_pie_chart(_("Ventes par catégorie"), sales_categories["labels"], sales_categories["values"]),
        pdf_pie_chart(_("Charges par catégorie"), expense_categories["labels"], expense_categories["values"]),
    ]
    category_drawings = [drawing for drawing in category_drawings if drawing is not None]
    if category_drawings:
        story.extend([Table([category_drawings]), Spacer(1, 8)])

    story.append(Paragraph(_("Séries des graphiques"), styles["Heading2"]))
    trend = context["chart_data"]["trend"]
    chart_rows = [[_("Date"), _("CA"), _("Achats"), _("Charges"), _("Gain brut"), _("Gain net")]]
    for index, label in enumerate(trend["labels"][:25]):
        chart_rows.append([
            label,
            trend["revenue"][index],
            trend["purchases"][index],
            trend["expenses"][index],
            trend["gross_profit"][index],
            trend["net_profit"][index],
        ])
    chart_table = Table(chart_rows, repeatRows=1)
    chart_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e9ecef")),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
    ]))
    story.append(chart_table)
    doc.build(story)
    return response
